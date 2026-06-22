from __future__ import annotations

from io import BytesIO
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import BinaryIO

from openpyxl import load_workbook

from guideline_pull.financial_data import fetch_financial_data

TICKER_ROWS = range(8, 18)

INFO_FIELD_TO_CELL = {
    "Name": "B",
    "Price per Share": "C",
    "Total Shares Outstanding": "D",
    "Total Debt": "G",
    "Interest Expense": "H",
    "Cash": "I",
    "Capital Expenditure": "J",
    "LTM Revenue": "K",
    "Gross Profit": "L",
    "Net Income": "M",
    "LTM EBIT": "N",
    "LTM EBITDA": "O",
    "Diluted EPS": "P",
    "Dividends Per Share": "Q",
    "Beta": "S",
    "Accounts Receivable": "W",
    "Inventory": "X",
    "Current Assets": "Y",
    "Accounts Payable": "Z",
    "Current Liabilities": "AA",
    "Total Assets": "AB",
    "Total Liabilities": "AC",
    "Depreciation": "AG",
    "OCF": "AI",
}


def _load_workbook_from_upload(uploaded_file: BinaryIO):
    uploaded_file.seek(0)
    return load_workbook(filename=BytesIO(uploaded_file.read()))


def extract_tickers_from_workbook(uploaded_file: BinaryIO) -> list[str]:
    """Read tickers from Info!A8:A17 in an uploaded workbook."""
    book = _load_workbook_from_upload(uploaded_file)
    if "Info" not in book.sheetnames:
        raise ValueError("Uploaded workbook must include a sheet named 'Info'.")

    sheet = book["Info"]
    tickers: list[str] = []
    for row in TICKER_ROWS:
        value = sheet[f"A{row}"].value
        if value:
            ticker = str(value).strip().upper()
            if ticker:
                tickers.append(ticker)
    return tickers


def _write_descriptions(book, data_by_ticker: dict[str, dict | None]) -> None:
    if "Descriptions" not in book.sheetnames:
        sheet = book.create_sheet("Descriptions")
    else:
        sheet = book["Descriptions"]

    sheet["A1"] = "Ticker"
    sheet["B1"] = "Company Name"
    sheet["C1"] = "Description"

    for row in range(2, sheet.max_row + 1):
        sheet[f"A{row}"] = None
        sheet[f"B{row}"] = None
        sheet[f"C{row}"] = None

    row = 2
    for ticker, values in data_by_ticker.items():
        if values:
            sheet[f"A{row}"] = ticker
            sheet[f"B{row}"] = values.get("Name", "N/A")
            sheet[f"C{row}"] = values.get("Description", "N/A")
            row += 1


def _write_info_sheet(book, data_by_ticker: dict[str, dict | None]) -> None:
    if "Info" not in book.sheetnames:
        raise ValueError("Uploaded workbook must include a sheet named 'Info'.")

    sheet = book["Info"]

    for row, ticker in zip(TICKER_ROWS, data_by_ticker.keys()):
        values = data_by_ticker[ticker]
        sheet[f"A{row}"] = ticker
        if not values:
            continue

        for field, column in INFO_FIELD_TO_CELL.items():
            sheet[f"{column}{row}"] = values.get(field, "N/A")


def fill_workbook(uploaded_file: BinaryIO) -> tuple[bytes, dict[str, dict | None], list[str]]:
    """Fill a guideline template workbook and return completed workbook bytes.

    Returns: workbook bytes, data_by_ticker, log messages.
    """
    book = _load_workbook_from_upload(uploaded_file)
    if "Info" not in book.sheetnames:
        raise ValueError("Uploaded workbook must include a sheet named 'Info'.")

    sheet = book["Info"]
    tickers = []
    for row in TICKER_ROWS:
        value = sheet[f"A{row}"].value
        if value:
            ticker = str(value).strip().upper()
            if ticker:
                tickers.append(ticker)

    if not tickers:
        raise ValueError("No tickers found in Info!A8:A17.")

    data_by_ticker: dict[str, dict | None] = {}
    logs: list[str] = []
    for ticker in tickers:
        data, messages = fetch_financial_data(ticker)
        logs.extend(messages)
        data_by_ticker[ticker] = data

    _write_info_sheet(book, data_by_ticker)
    _write_descriptions(book, data_by_ticker)

    # Instruct Excel to recalculate formula-driven sections when the file opens.
    try:
        book.calculation.fullCalcOnLoad = True
        book.calculation.forceFullCalc = True
    except Exception:
        pass

    with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        temp_path = Path(tmp.name)
    try:
        book.save(temp_path)
        output_bytes = temp_path.read_bytes()
    finally:
        temp_path.unlink(missing_ok=True)

    return output_bytes, data_by_ticker, logs
